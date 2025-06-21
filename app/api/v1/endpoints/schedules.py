from fastapi import APIRouter, Depends, HTTPException, status, Response
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_
from typing import List, Optional
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app.database import get_db
from app.models.schedule import Schedule
from app.models.subject import Subject
from app.models.teacher import Teacher
from app.models.class_type import ClassType
from app.models.classroom import Classroom
from app.schemas.schedule import ScheduleCreate, ScheduleUpdate, ScheduleResponse

router = APIRouter()


@router.post("/", response_model=ScheduleResponse, status_code=status.HTTP_201_CREATED)
def create_schedule(schedule: ScheduleCreate, db: Session = Depends(get_db)):
    """Crear un nuevo horario"""
    # Verificar que la asignatura existe
    subject = db.query(Subject).filter(Subject.id == schedule.subject_id).first()
    if not subject:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asignatura no encontrada"
        )
    
    # Verificar que el tipo de clase existe
    class_type = db.query(ClassType).filter(ClassType.id == schedule.class_type_id).first()
    if not class_type:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Tipo de clase no encontrado"
        )
    
    # Verificar que el aula existe
    classroom = db.query(Classroom).filter(Classroom.id == schedule.classroom_id).first()
    if not classroom:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Aula no encontrada"
        )
    
    # Verificar que el profesor existe
    teacher = db.query(Teacher).filter(Teacher.id == schedule.teacher_id).first()
    if not teacher:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Profesor no encontrado"
        )
    
    # Verificar conflictos de horario para el aula
    conflicting_schedule = db.query(Schedule).filter(
        and_(
            Schedule.classroom_id == schedule.classroom_id,
            Schedule.day_of_week == schedule.day_of_week,
            Schedule.semester == schedule.semester,
            Schedule.is_active == True,
            or_(
                and_(Schedule.start_time <= schedule.start_time, Schedule.end_time > schedule.start_time),
                and_(Schedule.start_time < schedule.end_time, Schedule.end_time >= schedule.end_time),
                and_(Schedule.start_time >= schedule.start_time, Schedule.end_time <= schedule.end_time)
            )
        )
    ).first()
    
    if conflicting_schedule:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Conflicto de horario: el aula ya está ocupada en este horario"
        )
    
    # Verificar conflictos de horario para el profesor
    conflicting_schedule = db.query(Schedule).filter(
        and_(
            Schedule.teacher_id == schedule.teacher_id,
            Schedule.day_of_week == schedule.day_of_week,
            Schedule.semester == schedule.semester,
            Schedule.is_active == True,
            or_(
                and_(Schedule.start_time <= schedule.start_time, Schedule.end_time > schedule.start_time),
                and_(Schedule.start_time < schedule.end_time, Schedule.end_time >= schedule.end_time),
                and_(Schedule.start_time >= schedule.start_time, Schedule.end_time <= schedule.end_time)
            )
        )
    ).first()
    
    if conflicting_schedule:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Conflicto de horario: el profesor ya tiene clase en este horario"
        )
    
    db_schedule = Schedule(**schedule.model_dump())
    db.add(db_schedule)
    db.commit()
    db.refresh(db_schedule)
    return db_schedule


@router.get("/", response_model=List[ScheduleResponse])
def get_schedules(
    skip: int = 0, 
    limit: int = 100, 
    semester: Optional[str] = None,
    teacher_id: Optional[int] = None,
    subject_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Obtener lista de horarios con filtros opcionales"""
    query = db.query(Schedule)
    
    if semester:
        query = query.filter(Schedule.semester == semester)
    
    if teacher_id:
        query = query.filter(Schedule.teacher_id == teacher_id)
    
    if subject_id:
        query = query.filter(Schedule.subject_id == subject_id)
    
    schedules = query.offset(skip).limit(limit).all()
    return schedules


@router.get("/{schedule_id}", response_model=ScheduleResponse)
def get_schedule(schedule_id: int, db: Session = Depends(get_db)):
    """Obtener un horario por ID"""
    schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if schedule is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Horario no encontrado"
        )
    return schedule


@router.put("/{schedule_id}", response_model=ScheduleResponse)
def update_schedule(schedule_id: int, schedule: ScheduleUpdate, db: Session = Depends(get_db)):
    """Actualizar un horario"""
    db_schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if db_schedule is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Horario no encontrado"
        )
    
    # Aquí se podrían agregar las mismas validaciones que en create_schedule
    # para verificar conflictos de horario
    
    update_data = schedule.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_schedule, field, value)
    
    db.commit()
    db.refresh(db_schedule)
    return db_schedule


@router.delete("/{schedule_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_schedule(schedule_id: int, db: Session = Depends(get_db)):
    """Eliminar un horario (soft delete)"""
    db_schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if db_schedule is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Horario no encontrado"
        )
    
    db_schedule.is_active = False
    db.commit()
    return None


@router.get("/export/weekly/{semester}")
def export_weekly_schedule(semester: str, db: Session = Depends(get_db)):
    """Exportar horario semanal a Excel"""
    schedules = db.query(Schedule).filter(
        and_(Schedule.semester == semester, Schedule.is_active == True)
    ).all()
    
    if not schedules:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No se encontraron horarios para este semestre"
        )
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Horario Semanal {semester}"
    
    # Configurar encabezados
    headers = ["Hora", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Horarios típicos (7:00 AM a 10:00 PM)
    time_slots = []
    start_time = datetime.strptime("07:00", "%H:%M")
    end_time = datetime.strptime("22:00", "%H:%M")
    
    current_time = start_time
    while current_time <= end_time:
        time_slots.append(current_time.strftime("%H:%M"))
        current_time += timedelta(hours=1)
    
    # Llenar horarios
    for row, time_slot in enumerate(time_slots, 2):
        ws.cell(row=row, column=1, value=time_slot)
        
        for day in range(6):  # 0-5 (Lunes a Sábado)
            # Buscar clases en este horario
            classes = [s for s in schedules if s.day_of_week == day]
            
            # Buscar clases que coincidan con este horario
            matching_classes = []
            for cls in classes:
                cls_start = cls.start_time.strftime("%H:%M")
                if cls_start == time_slot:
                    subject = db.query(Subject).filter(Subject.id == cls.subject_id).first()
                    teacher = db.query(Teacher).filter(Teacher.id == cls.teacher_id).first()
                    class_type = db.query(ClassType).filter(ClassType.id == cls.class_type_id).first()
                    classroom = db.query(Classroom).filter(Classroom.id == cls.classroom_id).first()
                    
                    class_info = f"{subject.acronym} - {class_type.acronym}\n{teacher.full_name}\n{classroom.code}"
                    matching_classes.append(class_info)
            
            if matching_classes:
                cell_value = "\n".join(matching_classes)
                cell = ws.cell(row=row, column=day + 2, value=cell_value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Ajustar ancho de columnas
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=horario_semanal_{semester}.xlsx"}
    )


@router.get("/export/teacher/{teacher_id}/{semester}")
def export_teacher_schedule(teacher_id: int, semester: str, db: Session = Depends(get_db)):
    """Exportar horario de un profesor específico a Excel"""
    # Verificar que el profesor existe
    teacher = db.query(Teacher).filter(Teacher.id == teacher_id).first()
    if not teacher:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Profesor no encontrado"
        )
    
    schedules = db.query(Schedule).filter(
        and_(
            Schedule.teacher_id == teacher_id,
            Schedule.semester == semester,
            Schedule.is_active == True
        )
    ).all()
    
    if not schedules:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No se encontraron horarios para este profesor en este semestre"
        )
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Horario {teacher.full_name}"
    
    # Configurar encabezados
    headers = ["Hora", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Horarios típicos
    time_slots = []
    start_time = datetime.strptime("07:00", "%H:%M")
    end_time = datetime.strptime("22:00", "%H:%M")
    
    current_time = start_time
    while current_time <= end_time:
        time_slots.append(current_time.strftime("%H:%M"))
        current_time += timedelta(hours=1)
    
    # Llenar horarios
    for row, time_slot in enumerate(time_slots, 2):
        ws.cell(row=row, column=1, value=time_slot)
        
        for day in range(6):
            # Buscar clases en este horario
            classes = [s for s in schedules if s.day_of_week == day]
            
            # Buscar clases que coincidan con este horario
            matching_classes = []
            for cls in classes:
                cls_start = cls.start_time.strftime("%H:%M")
                if cls_start == time_slot:
                    subject = db.query(Subject).filter(Subject.id == cls.subject_id).first()
                    class_type = db.query(ClassType).filter(ClassType.id == cls.class_type_id).first()
                    classroom = db.query(Classroom).filter(Classroom.id == cls.classroom_id).first()
                    
                    class_info = f"{subject.acronym} - {class_type.acronym}\n{classroom.code}"
                    matching_classes.append(class_info)
            
            if matching_classes:
                cell_value = "\n".join(matching_classes)
                cell = ws.cell(row=row, column=day + 2, value=cell_value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Ajustar ancho de columnas
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=horario_{teacher.full_name.replace(' ', '_')}_{semester}.xlsx"}
    ) 